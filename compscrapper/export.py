"""Exportación del resultado curado a Excel, con el formato del script original.

Solo se exportan las noticias marcadas como `mantener`. Los enlaces son
hipervínculos clickeables; la relevancia máxima se resalta en verde.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .models import ActorCandidato, Noticia

_HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HIGH_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
_LINK_FONT = Font(color="0563C1", underline="single")


def _write_sheet(ws, columns, rows, highlight=None):
    for col, (header, _key, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER
        ws.column_dimensions[cell.column_letter].width = width
    for row_idx, row in enumerate(rows, 2):
        for col, (_header, key, _width) in enumerate(columns, 1):
            value = row.get(key, "")
            if hasattr(value, "strftime"):
                value = value.strftime("%Y-%m-%d")
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = _BORDER
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            if key.startswith("url") and value:
                cell.hyperlink = value
                cell.font = _LINK_FONT
            if highlight and highlight(row):
                cell.fill = _HIGH_FILL
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = ws.dimensions


def _noticia_row(n: Noticia) -> dict:
    return {
        "relevancia": n.relevancia,
        "titulo": n.titulo,
        "fuente": n.fuente,
        "terminos": ", ".join(sorted(set(n.terminos))),
        "fecha_publicacion": n.fecha_publicacion,
        "url": n.url,
        "url_resaltado": n.url_resaltado,
        "consulta": n.consulta,
    }


def _write_parametros(ws, parametros: dict) -> None:
    """Hoja de procedencia: con qué se buscó, para que el resultado sea auditable."""
    ws.cell(row=1, column=1, value="Parámetro").font = _HEADER_FONT
    ws.cell(row=1, column=1).fill = _HEADER_FILL
    ws.cell(row=1, column=1).border = _BORDER
    ws.cell(row=1, column=2, value="Valor").font = _HEADER_FONT
    ws.cell(row=1, column=2).fill = _HEADER_FILL
    ws.cell(row=1, column=2).border = _BORDER
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 90
    for i, (clave, valor) in enumerate(parametros.items(), 2):
        if isinstance(valor, (list, tuple)):
            valor = "\n".join(str(v) for v in valor) if valor else "(ninguno)"
        ws.cell(row=i, column=1, value=str(clave)).border = _BORDER
        c = ws.cell(row=i, column=2, value=str(valor))
        c.border = _BORDER
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"


def export_excel(
    noticias: list[Noticia],
    actores: list[ActorCandidato],
    destino: str | Path,
    solo_mantenidas: bool = True,
    parametros: dict | None = None,
) -> Path:
    """Escribe el Excel y devuelve la ruta. Por defecto, solo las mantenidas.

    Si se pasan `parametros`, agrega una hoja "Parámetros" con la procedencia
    de la búsqueda (sector, consultas, ámbito, rango de fechas, fecha, versión).
    """
    filas = [n for n in noticias if (n.mantener or not solo_mantenidas)]
    wb = Workbook()

    ws = wb.active
    ws.title = "Noticias"
    _write_sheet(
        ws,
        [
            ("Relevancia (1-3)", "relevancia", 14),
            ("Título", "titulo", 60),
            ("Fuente", "fuente", 22),
            ("Términos hallados", "terminos", 35),
            ("Fecha publicación", "fecha_publicacion", 18),
            ("URL directa", "url", 45),
            ("URL con resaltado", "url_resaltado", 45),
            ("Consulta que la encontró", "consulta", 40),
        ],
        [_noticia_row(n) for n in filas],
        highlight=lambda r: r["relevancia"] == 3,
    )

    actores_filtrados = [a for a in actores if (a.mantener or not solo_mantenidas)]
    if actores_filtrados:
        ws2 = wb.create_sheet("Actores candidatos")
        _write_sheet(
            ws2,
            [
                ("Actor candidato", "candidato", 35),
                ("Menciones", "menciones", 12),
                ("Titular de ejemplo", "titular_ejemplo", 80),
            ],
            [
                {
                    "candidato": a.candidato,
                    "menciones": a.menciones,
                    "titular_ejemplo": a.titular_ejemplo,
                }
                for a in actores_filtrados
            ],
        )

    if parametros:
        _write_parametros(wb.create_sheet("Parámetros"), parametros)

    destino = Path(destino)
    wb.save(destino)
    return destino
