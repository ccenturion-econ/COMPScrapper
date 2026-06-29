"""Rango de fechas en consultas y hoja de procedencia en el Excel.

Ejecutable directo:  python tests/test_export.py
"""

from __future__ import annotations

import sys
import tempfile
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
from compscrapper.engine import date_clause, expand_queries  # noqa: E402
from compscrapper.export import export_excel  # noqa: E402
from compscrapper.models import Noticia  # noqa: E402


def test_date_clause():
    assert date_clause() == ""
    assert date_clause(date(2025, 1, 1)) == "after:2025-01-01"
    assert date_clause(hasta=date(2025, 6, 30)) == "before:2025-06-30"
    assert date_clause(date(2025, 1, 1), date(2025, 6, 30)) == "after:2025-01-01 before:2025-06-30"


def test_expand_con_fechas():
    # Ámbito nacional ancla en el grupo país; las fechas se aplican al final.
    out = expand_queries(["carne"], [], desde=date(2025, 1, 1), hasta=date(2025, 6, 30))
    assert len(out) == 1
    assert out[0].startswith("carne (Paraguay OR")
    assert out[0].endswith(") after:2025-01-01 before:2025-06-30")


def test_export_con_procedencia(tmp_path: Path):
    from openpyxl import load_workbook
    noticias = [
        Noticia(3, "Titular A", "Diario", ["carne"], None, "https://x.com/a", "", "q1", True),
        Noticia(2, "Titular B", "Diario", ["precio"], None, "https://x.com/b", "", "q2", False),
    ]
    destino = tmp_path / "out.xlsx"
    export_excel(noticias, [], destino, solo_mantenidas=True,
                 parametros={"Sector": "Carne", "Ámbito geográfico": ["Nacional"]})
    wb = load_workbook(destino)
    assert "Parámetros" in wb.sheetnames
    # solo la noticia mantenida (A) va en la hoja Noticias.
    ws = wb["Noticias"]
    titulos = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]
    assert titulos == ["Titular A"]


if __name__ == "__main__":
    test_date_clause(); print("OK  test_date_clause")
    test_expand_con_fechas(); print("OK  test_expand_con_fechas")
    with tempfile.TemporaryDirectory() as d:
        test_export_con_procedencia(Path(d)); print("OK  test_export_con_procedencia")
    print("\n3 pruebas de export/fechas pasaron.")
